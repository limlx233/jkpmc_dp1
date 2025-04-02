import re
import calendar
import pandas as pd
from io import BytesIO
import streamlit as st
import dataprocess as dp  # 根据实际处理需求编写的数据处理模块
from datetime import date
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.formatting.rule import DataBarRule, FormulaRule


# 将 Pandas DataFrame 对象转换为 Excel 文件格式的字节流
@st.cache_resource
def to_excel(df_s11, df_s12, df_s2, df_s3, df_s4, df_s5, df2,
            sheet_name1='正常品种销售-口腔', sheet_name2='正常品种销售-洗护', sheet_name3='电商',
            sheet_name4='促销品&非卖', sheet_name5='拓展部', sheet_name6='齿说', sheet_names='异常类别定义'):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df_all = pd.concat([df_s11, df_s12, df_s2, df_s3, df_s4, df_s5])
    # 将多个 DataFrame 存入字典
    df_dict = {
        '汇总':df_all, # 新增明细汇总 
        '口腔数据': df_s11,
        '洗护数据': df_s12,
        '电商数据': df_s2,
        '促销品&非卖数据': df_s3,
        '拓展部数据': df_s4,
        '齿说数据': df_s5
    }
    # 创建一个新的工作表，用于集中写入所有 DataFrame
    combined_sheet_name = '汇总数据'
    worksheet_combined = writer.book.create_sheet(combined_sheet_name)
    # 初始化当前写入行
    current_row = 1
    # 循环处理字典中的每个 DataFrame
    for df_name, df in df_dict.items():
        # 按分类分组并进行聚合操作
        result = df.groupby('分类',observed=False).agg({
            '产品编码': 'nunique', # 'count'
            '库存总件数': 'sum'
        }).reset_index()
        # 重命名列名
        result.columns = ['分类', 'SKU个数', '件数']
        # 写入来源描述
        worksheet_combined.cell(row=current_row, column=1, value=f'数据来源: {df_name}')
        current_row += 1
        # 写入列名
        for col_index, col_name in enumerate(result.columns, start=1):
            worksheet_combined.cell(row=current_row, column=col_index, value=col_name)
        current_row += 1
        # 写入数据
        for row in result.values.tolist():
            for col_index, value in enumerate(row, start=1):
                worksheet_combined.cell(row=current_row, column=col_index, value=value)
            current_row += 1
        # 空一行分隔不同 DataFrame 的结果
        current_row += 1

    df2.to_excel(writer, index=False, header=False, sheet_name=sheet_names)
    df_s11.to_excel(writer, index=False, sheet_name=sheet_name1)
    df_s12.to_excel(writer, index=False, sheet_name=sheet_name2)
    df_s2.to_excel(writer, index=False, sheet_name=sheet_name3)
    df_s3.to_excel(writer, index=False, sheet_name=sheet_name4)
    df_s4.to_excel(writer, index=False, sheet_name=sheet_name5)
    df_s5.to_excel(writer, index=False, sheet_name=sheet_name6)
    # 获取工作簿和工作表
    workbook = writer.book
    worksheet1 = writer.sheets[sheet_names]
    worksheet2 = writer.sheets[sheet_name1]
    worksheet3 = writer.sheets[sheet_name2]
    worksheet4 = writer.sheets[sheet_name3]
    worksheet5 = writer.sheets[sheet_name4]
    worksheet6 = writer.sheets[sheet_name5]
    worksheet7 = writer.sheets[sheet_name6]
    worksheet_combined = writer.sheets[combined_sheet_name]

    # 确保百分比样式只被创建一次
    if "percentage_style" not in workbook.named_styles:
        percentage_style = NamedStyle(name="percentage_style", number_format='0.00%')
        workbook.add_named_style(percentage_style)
    else:
        percentage_style = workbook.named_styles["percentage_style"]

    # 调用格式设置函数
    set_description_sheet_format(worksheet1, df2)
    set_material_sheet_format(worksheet2, df_s11, percentage_style)
    set_material_sheet_format(worksheet3, df_s12, percentage_style)
    set_material_sheet_format(worksheet4, df_s2, percentage_style)
    set_material_sheet_format(worksheet5, df_s3, percentage_style)
    set_material_sheet_format(worksheet6, df_s4, percentage_style)
    set_material_sheet_format(worksheet7, df_s5, percentage_style)
    add_data_bar_rule(worksheet2, start_row=2, end_row=len(df1) + 1, column='E')
    add_data_bar_rule(worksheet3, start_row=2, end_row=len(df1) + 1, column='E')
    add_data_bar_rule(worksheet4, start_row=2, end_row=len(df1) + 1, column='E')
    add_data_bar_rule(worksheet5, start_row=2, end_row=len(df1) + 1, column='E')
    add_data_bar_rule(worksheet6, start_row=2, end_row=len(df1) + 1, column='E')
    add_data_bar_rule(worksheet7, start_row=2, end_row=len(df1) + 1, column='E')
    writer.close()
    processed_data = output.getvalue()
    return processed_data



def add_data_bar_rule(worksheet, start_row, end_row, column, color="c00000"):
    # 范围字符串
    range_str = f'{column}{start_row}:{column}{end_row}'

    # 添加条件格式规则，排除负数
    negative_rule = FormulaRule(formula=[f'AND({column}{start_row}<0)'], stopIfTrue=True)
    worksheet.conditional_formatting.add(range_str, negative_rule)

    # 添加数据条规则
    data_bar_rule = DataBarRule(
        start_type='num', start_value=0,
        end_type='max',
        color=color
    )
    worksheet.conditional_formatting.add(range_str, data_bar_rule)

# 示例用法
# add_data_bar_rule(worksheet, start_row=2, end_row=len(df1) + 1, column='E')

def set_material_sheet_format(worksheet, df1, percentage_style):
    # 设置列宽
    for idx, column in enumerate(worksheet.columns, start=1):
        column_letter = column[0].column_letter
        if idx == 1:
            worksheet.column_dimensions[column_letter].width = 16
        elif idx == 5:
            worksheet.column_dimensions[column_letter].width = 22
        elif idx == 7:
            worksheet.column_dimensions[column_letter].width = 30
        else:
            worksheet.column_dimensions[column_letter].width = 16

    # 应用百分比格式到指定列
    column_name = '%(剩余效期/总效期)'
    column_index = df1.columns.get_loc(column_name) + 1  # +1 是因为 Excel 列索引从 1 开始
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            cell.style = percentage_style

    # 定义将 RGB 颜色值转换为十六进制颜色代码的函数
    def rgb_to_hex(r, g, b):
        return '{:02x}{:02x}{:02x}'.format(r, g, b)

    # 定义不同列所需的颜色
    colors = {
        1: rgb_to_hex(226, 107, 10),
        2: rgb_to_hex(49, 134, 155),
        3: rgb_to_hex(226, 107, 10),
        4: rgb_to_hex(226, 107, 10),
        5: rgb_to_hex(118, 147, 60),
        6: rgb_to_hex(118, 147, 60)
    }

    # 默认颜色
    default_color = "346c9c"

    # 遍历第一行的每个单元格
    for col_index, cell in enumerate(worksheet[1], start=1):
        # 设置字体为加粗，颜色为白色
        cell.font = Font(bold=True, color="FFFFFF")
        # 根据列索引获取相应的颜色，如果没有指定则使用默认颜色
        fill_color = colors.get(col_index, default_color)
        # 设置填充颜色
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        # 设置单元格内容居中对齐
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if cell.column == 7 or cell.column == 14:  # G列是第7列
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")



def set_description_sheet_format(worksheet, df2):
    # 设置列宽
    worksheet.column_dimensions['A'].width = 22
    worksheet.column_dimensions['B'].width = 108
    # 设置行高
    for row in range(1, 12):  # 1~11行行高设置为36
        worksheet.row_dimensions[row].height = 36
    # 设置标题样式
    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 设置剩余行的标题样式
    for row in range(2, worksheet.max_row + 1):  # 从第二行开始
        for cell in worksheet[row]:
            cell.font = Font(size=14, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    # 合并A1和B1
    worksheet.merge_cells('A1:B1')

    # 设置合并单元格的对齐方式
    worksheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # 保证样式设置不会被合并操作覆盖
    for row in worksheet['A1:B1']:
        for cell in row:
            cell.font = Font(bold=True, size=16)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置A2:A8单元格的样式
    for row in worksheet['A2:A8']:
        for cell in row:
            cell.font = Font(size=14, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # 设置C2:C8单元格的样式
    for row in worksheet['C2:C8']:
        for cell in row:
            cell.font = Font(size=14, color="000000", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

# 页面设置
st.set_page_config(page_title="数据处理工具", page_icon=":material/home:", layout='centered')

with st.container(border=True):
    st.header('成品库存异常情况数据处理', divider="rainbow")
    st.subheader('月末日期选择', divider='grey')

    def get_last_day_of_previous_month():
        today = date.today()
        if today.month == 1:
            year = today.year - 1
            month = 12
        else:
            year = today.year
            month = today.month - 1
        _, last_day = calendar.monthrange(year, month)
        return date(year, month, last_day)

    # 默认日期为上月最后一天
    default_date = get_last_day_of_previous_month()
    # 日期输入控件
    date_value = st.date_input(label="请选择日期,(默认为上月的最后一天)", value=default_date)
    # st.write(date_value)
    st.subheader('1.库存数据文件上传', divider='grey')
    uploaded_file1 = st.file_uploader(label="请选择库存数据Excel文件(.xlsx格式)上传", accept_multiple_files=False, type=["xlsx"])
    st.subheader('2.月末呆滞数据文件上传', divider='grey')
    uploaded_file2 = st.file_uploader(label="请选择呆滞数据Excel文件(.xlsx格式)上传", accept_multiple_files=False, type=["xlsx"])
    col1, col2 = st.columns(2)

    with col1:
        if st.button(label="数据处理", type="primary", key="data_process"):
            if uploaded_file1 and uploaded_file2:
                # 读取文件并缓存
                df1 = pd.read_excel(uploaded_file1, header=1)  # 库存信息
                df2 = pd.read_excel(uploaded_file2, header=0)  # 呆滞数据文件
                df_res, df2_res = dp.read_data(df1, df2)
                df_res = dp.calculate_expiry(df_res, date_value)
                df_res = dp.expiry_classification(df_res)
                df_res = dp.merge_and_mark(df_res, df2_res)
                df_res = dp.classify_items(df_res)
                df_res = dp.filter_and_calculate(df_res)
                df_res = dp.sort_and_filter(df_res)
                df_s11, df_s12, df_s2, df_s3, df_s4, df_s5 = dp.filter_special_cases(df_res)
                df2 = dp.generate_description_df()
                
                excel_file = to_excel(df_s11, df_s12, df_s2, df_s3, df_s4, df_s5, df2)
                st.session_state.excel_file = excel_file
            else:
                st.info("请先上传数据文件!")
    with col2:
        if 'excel_file' in st.session_state:
            st.download_button(
                label="下载文件",
                data=st.session_state.excel_file,
                type="primary",
                file_name="产成品月末库存异常情况.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("请先上传数据并进行数据处理。")